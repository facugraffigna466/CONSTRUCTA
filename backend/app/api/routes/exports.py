"""
Export endpoint — generates an .xlsx file with all tasks for a given obra.

GET /api/v1/exports/obras/{obra_id}/excel
"""
import io
from datetime import date

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.deps import CurrentUser, CurrentUserId, DbSession
from app.services.obra_service import ObraService
from app.repositories.responsible import ResponsibleRepository
from app.repositories.task import TaskRepository

router = APIRouter(prefix="/exports", tags=["exports"])

# ─── Status labels ────────────────────────────────────────────────────────────

STATUS_LABEL = {
    "pendiente":   "Pendiente",
    "en_progreso": "En progreso",
    "bloqueada":   "Bloqueada",
    "completada":  "Completada",
    "cancelada":   "Cancelada",
}

STATUS_COLOR = {
    "pendiente":   "DBEAFE",   # blue-100
    "en_progreso": "FEF3C7",   # amber-100
    "bloqueada":   "FEE2E2",   # red-100
    "completada":  "D1FAE5",   # green-100
    "cancelada":   "F3F4F6",   # gray-100
}

HEADER_FILL  = PatternFill("solid", fgColor="1B1B1A")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
SUBHEAD_FILL = PatternFill("solid", fgColor="F5F0E8")
SUBHEAD_FONT = Font(bold=True, color="5B6770", size=9)

COLUMNS = [
    ("N°",          6),
    ("Tarea",       42),
    ("Hito",        7),
    ("Responsable", 24),
    ("Inicio",      13),
    ("Vencimiento", 13),
    ("Duración (d)", 13),
    ("Estado",      14),
    ("% Avance",    11),
]


def _fmt_date(d: date | None) -> str:
    return d.strftime("%d/%m/%Y") if d else ""


def _duration(start: date | None, due: date | None) -> int | str:
    if start and due:
        return (due - start).days
    return ""


@router.get("/obras/{obra_id}/excel")
async def export_tasks_excel(
    obra_id: int,
    db: DbSession,
    current_user: CurrentUser,
):
    await ObraService(db).get_or_raise(obra_id, tenant_id=current_user.tenant_id)
    tasks = await TaskRepository(db).list_by_obra(obra_id)
    if not tasks:
        raise HTTPException(404, "No hay tareas para exportar.")

    # Build responsible lookup
    resp_repo  = ResponsibleRepository(db)
    resp_ids   = {t.responsible_id for t in tasks if t.responsible_id}
    resp_names: dict[int, str] = {}
    for rid in resp_ids:
        r = await resp_repo.get(rid)
        if r:
            resp_names[rid] = r.full_name

    wb = Workbook()
    ws = wb.active
    ws.title = "Tareas"

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # ── Task rows ─────────────────────────────────────────────────────────────
    for row_idx, task in enumerate(tasks, start=2):
        status_str = task.status.value if hasattr(task.status, "value") else str(task.status)
        fill_color = STATUS_COLOR.get(status_str, "FFFFFF")
        row_fill   = PatternFill("solid", fgColor=fill_color)

        values = [
            row_idx - 1,
            task.title,
            "◆" if task.is_milestone else "",
            resp_names.get(task.responsible_id, "") if task.responsible_id else "",
            _fmt_date(task.start_date),
            _fmt_date(task.due_date),
            _duration(task.start_date, task.due_date),
            STATUS_LABEL.get(status_str, status_str),
            task.estimated_progress,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center")
            cell.font = Font(size=10)

        ws.row_dimensions[row_idx].height = 18

    # Right-align numeric columns (N°, Duración, % Avance)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col in (1, 7, 9):
            row[col - 1].alignment = Alignment(horizontal="right", vertical="center")

    # ── Stream response ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="obra_{obra_id}_tareas.xlsx"'},
    )


# ─── Template de importación ──────────────────────────────────────────────────

@router.get("/template-excel")
async def download_template_excel(_: CurrentUserId):
    """Plantilla .xlsx vacía con las columnas que espera el import + instrucciones."""
    wb = Workbook()

    # Hoja de tareas
    ws = wb.active
    ws.title = "Tareas"
    template_cols = [
        ("Nombre",       42),
        ("Inicio",       14),
        ("Fin",          14),
        ("Responsable",  24),
        ("Predecesoras", 14),
    ]
    for idx, (label, width) in enumerate(template_cols, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 22

    example_rows = [
        ("Excavación y movimiento de suelos", "2026-07-01", "2026-07-10", "Juan Pérez", ""),
        ("Hormigón de fundaciones",           "2026-07-11", "2026-07-25", "Juan Pérez", "1"),
        ("Mampostería planta baja",           "2026-07-26", "2026-08-15", "",           "2"),
    ]
    example_fill = PatternFill("solid", fgColor="F5F0E8")
    for r, values in enumerate(example_rows, start=2):
        for c, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.fill = example_fill
            cell.font = Font(size=10, italic=True, color="8E97A0")

    # Hoja de instrucciones
    inst = wb.create_sheet("Instrucciones")
    inst.column_dimensions["A"].width = 95
    lines = [
        "CÓMO IMPORTAR TAREAS A CONSTRUCTA",
        "",
        "1. Completá la hoja 'Tareas'. Las filas de ejemplo (en gris) podés borrarlas o pisarlas.",
        "2. Columnas:",
        "   • Nombre — obligatorio. Título de la tarea.",
        "   • Inicio / Fin — opcional. Formato AAAA-MM-DD o DD/MM/AAAA.",
        "   • Responsable — opcional. Si coincide con alguien del equipo de la obra, se asigna solo.",
        "   • Predecesoras — opcional. Número de FILA de la tarea de la que depende (ej: 2).",
        "3. Guardá el archivo y subilo desde el botón 'Importar' en el tab Tareas de tu obra.",
        "4. Vas a ver una vista previa antes de confirmar — nada se crea hasta que confirmes.",
        "",
        "También podés importar directo un .xml exportado desde MS Project",
        "(Archivo → Guardar como → XML): se respetan subtareas, hitos, dependencias",
        "FS/SS/FF/SF con lag y los recursos asignados.",
    ]
    for r, text in enumerate(lines, start=1):
        cell = inst.cell(row=r, column=1, value=text)
        if r == 1:
            cell.font = Font(bold=True, size=12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="constructa_plantilla_tareas.xlsx"'},
    )


# ─── Presupuesto por obra ─────────────────────────────────────────────────────

@router.get("/obras/{obra_id}/presupuesto-excel")
async def export_presupuesto_excel(obra_id: int, db: DbSession, _: CurrentUserId):
    """Excel del presupuesto: materiales por tarea con subtotales y totales."""
    from sqlalchemy import select
    from app.models.obra import Obra
    from app.models.supplier import Supplier
    from app.models.task import Task
    from app.models.task_material import TaskMaterial

    obra = (await db.execute(select(Obra).where(Obra.id == obra_id))).scalar_one_or_none()
    if not obra:
        raise HTTPException(404, "Obra no encontrada")

    result = await db.execute(
        select(TaskMaterial, Task.title, Supplier.name)
        .join(Task, TaskMaterial.task_id == Task.id)
        .outerjoin(Supplier, TaskMaterial.supplier_id == Supplier.id)
        .where(Task.obra_id == obra_id)
        .order_by(Task.order_index, Task.id, TaskMaterial.created_at)
    )
    rows = result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"

    cols = [
        ("Tarea", 36), ("Ítem", 32), ("Cantidad", 11), ("Unidad", 10),
        ("Precio unit.", 13), ("Subtotal", 14), ("Estado", 12), ("Proveedor", 22),
    ]
    for idx, (label, width) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 22

    MATERIAL_STATUS_LABEL = {"pendiente": "Pendiente", "pedido": "Pedido", "recibido": "Recibido"}
    MATERIAL_STATUS_COLOR = {"pendiente": "DBEAFE", "pedido": "FEF3C7", "recibido": "D1FAE5"}

    total_estimado = total_recibido = 0.0
    r = 2
    for material, task_title, supplier_name in rows:
        qty = float(material.quantity) if material.quantity is not None else None
        price = float(material.unit_price) if material.unit_price is not None else None
        subtotal = (qty or 0) * (price or 0)
        total_estimado += subtotal
        if material.status == "recibido":
            total_recibido += subtotal
        fill = PatternFill("solid", fgColor=MATERIAL_STATUS_COLOR.get(material.status, "FFFFFF"))
        values = [task_title, material.name, qty, material.unit or "",
                  price, subtotal, MATERIAL_STATUS_LABEL.get(material.status, material.status),
                  supplier_name or ""]
        for c, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical="center")
            if c == 7:
                cell.fill = fill
        for c in (3, 5, 6):
            ws.cell(row=r, column=c).number_format = "#,##0.00"
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="right", vertical="center")
        r += 1

    # Totales
    r += 1
    ws.cell(row=r, column=5, value="Total estimado").font = Font(bold=True, size=10)
    tot = ws.cell(row=r, column=6, value=total_estimado)
    tot.font = Font(bold=True, size=10)
    tot.number_format = "#,##0.00"
    r += 1
    ws.cell(row=r, column=5, value="Gasto real (recibido)").font = Font(bold=True, size=10, color="1F8A5B")
    tot2 = ws.cell(row=r, column=6, value=total_recibido)
    tot2.font = Font(bold=True, size=10, color="1F8A5B")
    tot2.number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="obra_{obra_id}_presupuesto.xlsx"'},
    )
