"""
Excel / CSV / MS Project XML import service.

Parses an uploaded file and returns a preview of rows to be confirmed.
Supports .xlsx (openpyxl), .csv (stdlib csv) and MS Project .xml
(xml.etree.ElementTree, formato documentado por Microsoft).
Column detection is automatic, case-insensitive, Spanish and English.
"""
import csv
import io
import xml.etree.ElementTree as ET
from datetime import date

import openpyxl

from app.schemas.imports import ImportDepLink, ImportPreview, ImportPreviewRow

# ── Column alias maps ─────────────────────────────────────────────────────────

_TITLE_ALIASES       = {"nombre", "name", "task name", "tarea", "título", "titulo", "actividad", "task"}
_START_ALIASES       = {"comienzo", "start", "inicio", "fecha inicio", "start date", "fecha de inicio"}
_END_ALIASES         = {"fin", "finish", "vencimiento", "fecha fin", "end", "end date", "due date", "fecha de vencimiento"}
_RESPONSIBLE_ALIASES = {"recursos", "resource names", "responsable", "asignado", "assigned to", "recurso", "resource"}
_PREDECESSOR_ALIASES = {"predecesoras", "predecessors", "dependencia", "depends on", "predecesora", "pred."}

_ALIAS_GROUPS: dict[str, set[str]] = {
    "title":        _TITLE_ALIASES,
    "start_date":   _START_ALIASES,
    "due_date":     _END_ALIASES,
    "responsible":  _RESPONSIBLE_ALIASES,
    "predecessors": _PREDECESSOR_ALIASES,
}


def _detect_column_map(headers: list[str]) -> dict[str, int | None]:
    normalized = [h.strip().lower() for h in headers]
    result: dict[str, int | None] = {}
    for field, aliases in _ALIAS_GROUPS.items():
        idx = next((i for i, h in enumerate(normalized) if h in aliases), None)
        result[field] = idx
    return result


def _parse_date_cell(raw: object) -> date | None:
    """Parse a cell value from openpyxl (may already be a date/datetime) or string."""
    if raw is None:
        return None
    if isinstance(raw, date):
        return raw if isinstance(raw, date) and not hasattr(raw, "hour") else raw.date()  # type: ignore[attr-defined]
    s = str(raw).strip()
    if not s:
        return None
    # ISO: YYYY-MM-DD
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            from datetime import datetime as _dt
            return _dt.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_predecessor(raw: object) -> int | None:
    if raw is None:
        return None
    s = str(raw).strip()
    import re
    m = re.search(r"\d+", s)
    if not m:
        return None
    row_num = int(m.group())
    return row_num - 1  # 1-based → 0-based


def _rows_from_sheet(sheet) -> tuple[list[str], list[list[object]]]:
    """Return (headers, data_rows) from an openpyxl sheet."""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(c or "") for c in rows[0]]
    data = [list(r) for r in rows[1:] if any(c is not None for c in r)]
    return headers, data


def _rows_from_csv(content: bytes) -> tuple[list[str], list[list[str]]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


# ── MS Project XML ────────────────────────────────────────────────────────────

# Tipos de dependencia en el XML de MS Project: 0=FF, 1=FS (default), 2=SF, 3=SS
_MSP_DEP_TYPE = {0: "FF", 1: "FS", 2: "SF", 3: "SS"}


def _strip_ns(tag: str) -> str:
    """'{http://schemas.microsoft.com/project}Task' → 'Task'"""
    return tag.rsplit("}", 1)[-1]


def _find_text(el: ET.Element, name: str) -> str | None:
    for child in el:
        if _strip_ns(child.tag) == name:
            return child.text
    return None


def _findall(el: ET.Element, name: str) -> list[ET.Element]:
    return [c for c in el if _strip_ns(c.tag) == name]


def _parse_msp_date(raw: str | None) -> date | None:
    if not raw:
        return None
    try:
        return date.fromisoformat(raw.strip()[:10])
    except ValueError:
        return None


def _parse_msp_lag(link: ET.Element) -> int:
    """LinkLag viene en décimas de minuto (4800 = 1 día de 8h).
    Algunos exports usan <Lag> en minutos (480 = 1 día)."""
    raw = _find_text(link, "LinkLag")
    if raw is not None:
        try:
            return round(int(raw) / 4800)
        except ValueError:
            return 0
    raw = _find_text(link, "Lag")
    if raw is not None:
        try:
            return round(int(raw) / 480)
        except ValueError:
            return 0
    return 0


def is_msproject_xml(file_bytes: bytes) -> bool:
    head = file_bytes[:200].lstrip()
    return head.startswith(b"<?xml") or head.startswith(b"<Project")


def parse_msproject_xml(file_bytes: bytes) -> ImportPreview:
    """Parsea el XML de MS Project: tareas + WBS + dependencias + recursos."""
    root = ET.fromstring(file_bytes)
    if _strip_ns(root.tag) != "Project":
        raise ValueError("El XML no tiene el formato de MS Project (falta el nodo <Project>).")

    # Recursos: UID → nombre
    resource_names: dict[str, str] = {}
    for resources in _findall(root, "Resources"):
        for res in _findall(resources, "Resource"):
            uid = _find_text(res, "UID")
            name = (_find_text(res, "Name") or "").strip()
            if uid is not None and name:
                resource_names[uid] = name

    # Asignaciones: TaskUID → primer recurso asignado
    task_resource: dict[str, str] = {}
    for assignments in _findall(root, "Assignments"):
        for asg in _findall(assignments, "Assignment"):
            t_uid = _find_text(asg, "TaskUID")
            r_uid = _find_text(asg, "ResourceUID")
            if t_uid and r_uid and r_uid in resource_names and t_uid not in task_resource:
                task_resource[t_uid] = resource_names[r_uid]

    # Tareas
    preview_rows: list[ImportPreviewRow] = []
    uid_to_row: dict[str, int] = {}
    raw_tasks: list[dict] = []

    for tasks_el in _findall(root, "Tasks"):
        for task_el in _findall(tasks_el, "Task"):
            uid = _find_text(task_el, "UID")
            name = (_find_text(task_el, "Name") or "").strip()
            outline_raw = _find_text(task_el, "OutlineLevel")
            try:
                outline = int(outline_raw) if outline_raw is not None else 1
            except ValueError:
                outline = 1
            # UID 0 / OutlineLevel 0 es la fila-resumen del proyecto — no se importa
            if uid == "0" or outline == 0:
                continue
            if not name:
                continue

            deps: list[dict] = []
            for link in _findall(task_el, "PredecessorLink"):
                pred_uid = _find_text(link, "PredecessorUID")
                if not pred_uid or pred_uid == "0":
                    continue
                try:
                    type_code = int(_find_text(link, "Type") or "1")
                except ValueError:
                    type_code = 1
                deps.append({
                    "pred_uid": pred_uid,
                    "dependency_type": _MSP_DEP_TYPE.get(type_code, "FS"),
                    "lag_days": _parse_msp_lag(link),
                })

            raw_tasks.append({
                "uid": uid,
                "name": name,
                "outline": outline,
                "start": _parse_msp_date(_find_text(task_el, "Start")),
                "finish": _parse_msp_date(_find_text(task_el, "Finish")),
                "milestone": (_find_text(task_el, "Milestone") or "0").strip() == "1",
                "parent_uid": _find_text(task_el, "ParentTaskUID"),
                "deps": deps,
            })

    # Índices de fila por UID (orden del archivo = orden de outline)
    for i, t in enumerate(raw_tasks):
        if t["uid"] is not None:
            uid_to_row[t["uid"]] = i

    # Jerarquía WBS: ParentTaskUID si existe; si no, stack por OutlineLevel
    warnings = 0
    outline_stack: list[tuple[int, int]] = []  # (outline_level, row_index)
    for i, t in enumerate(raw_tasks):
        parent_row: int | None = None
        if t["parent_uid"] and t["parent_uid"] != "0" and t["parent_uid"] in uid_to_row:
            parent_row = uid_to_row[t["parent_uid"]]
        else:
            while outline_stack and outline_stack[-1][0] >= t["outline"]:
                outline_stack.pop()
            if outline_stack:
                parent_row = outline_stack[-1][1]
        outline_stack.append((t["outline"], i))

        dep_links: list[ImportDepLink] = []
        for d in t["deps"]:
            row = uid_to_row.get(d["pred_uid"])
            if row is not None:
                dep_links.append(ImportDepLink(
                    row=row,
                    dependency_type=d["dependency_type"],
                    lag_days=d["lag_days"],
                ))

        row_warnings: list[str] = []
        if t["start"] and t["finish"] and t["finish"] < t["start"]:
            row_warnings.append("Fin anterior al inicio")
        warning_str = "; ".join(row_warnings) if row_warnings else None
        if warning_str:
            warnings += 1

        preview_rows.append(ImportPreviewRow(
            row_index=i,
            title=t["name"],
            start_date=t["start"],
            due_date=t["finish"],
            responsible_name=task_resource.get(t["uid"]),
            dependency_links=dep_links or None,
            parent_row=parent_row,
            is_milestone=t["milestone"],
            warning=warning_str,
        ))

    return ImportPreview(
        rows=preview_rows,
        column_map={
            "title": "Name", "start_date": "Start", "due_date": "Finish",
            "responsible": "Resources/Assignments", "predecessors": "PredecessorLink",
        },
        total_rows=len(preview_rows),
        warnings=warnings,
        errors=0,
        source="msproject",
    )


def parse_excel(
    file_bytes: bytes, mime: str, column_overrides: dict[str, str] | None = None
) -> ImportPreview:
    """column_overrides: {campo: nombre_de_encabezado} — remapeo manual cuando
    la detección automática falla (solo aplica a Excel/CSV, no a MS Project XML)."""
    if is_msproject_xml(file_bytes) or mime in ("text/xml", "application/xml"):
        return parse_msproject_xml(file_bytes)
    if mime == "text/csv" or mime == "application/csv":
        headers, data_rows = _rows_from_csv(file_bytes)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        headers, data_rows = _rows_from_sheet(ws)

    col_map = _detect_column_map(headers)
    if column_overrides:
        norm = [str(h).strip().lower() for h in headers]
        for field, header_name in column_overrides.items():
            if field not in col_map:
                continue
            if not header_name:
                col_map[field] = None
                continue
            try:
                col_map[field] = norm.index(header_name.strip().lower())
            except ValueError:
                pass
    col_names: dict[str, str] = {
        field: headers[idx] for field, idx in col_map.items() if idx is not None
    }

    preview_rows: list[ImportPreviewRow] = []
    warnings = 0
    errors = 0

    for i, row in enumerate(data_rows):
        def cell(field: str) -> object:
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        title = str(cell("title") or "").strip()
        if not title:
            errors += 1
            preview_rows.append(ImportPreviewRow(
                row_index=i, title="(sin título)", error="Fila sin título — no se importará."
            ))
            continue

        start_date = _parse_date_cell(cell("start_date"))
        due_date   = _parse_date_cell(cell("due_date"))
        responsible_name = str(cell("responsible") or "").strip() or None
        depends_on_row   = _parse_predecessor(cell("predecessors"))

        row_warnings: list[str] = []
        if cell("start_date") and start_date is None:
            row_warnings.append("Fecha de inicio inválida")
        if cell("due_date") and due_date is None:
            row_warnings.append("Fecha de vencimiento inválida")
        if start_date and due_date and due_date < start_date:
            row_warnings.append("Vencimiento anterior al inicio")

        warning_str = "; ".join(row_warnings) if row_warnings else None
        if warning_str:
            warnings += 1

        preview_rows.append(ImportPreviewRow(
            row_index=i,
            title=title,
            start_date=start_date,
            due_date=due_date,
            responsible_name=responsible_name,
            depends_on_row=depends_on_row,
            warning=warning_str,
        ))

    return ImportPreview(
        rows=preview_rows,
        column_map=col_names,
        headers=[str(h) for h in headers if str(h).strip()],
        total_rows=len(preview_rows),
        warnings=warnings,
        errors=errors,
    )


# ── AI-powered column detection ──────────────────────────────────────────────

import hashlib
import json as _json

# Monthly AI call limits per plan name (None = unlimited)
_AI_MONTHLY_LIMITS: dict[str, int | None] = {
    "basico": 10,
    "pro": 50,
    "enterprise": None,
}
_DEFAULT_LIMIT = 5  # users without a plan

_FIELD_DESCRIPTIONS = {
    "title":        "Nombre o título de la tarea/actividad (requerido)",
    "start_date":   "Fecha de inicio de la tarea",
    "due_date":     "Fecha de fin, vencimiento o finalización",
    "responsible":  "Persona responsable o asignada a la tarea",
    "predecessors": "Tareas predecesoras o dependencias (números de fila o IDs)",
}


def _headers_hash(headers: list[str]) -> str:
    key = "|".join(h.strip().lower() for h in headers)
    return hashlib.sha256(key.encode()).hexdigest()


async def detect_column_mapping_ai(
    file_bytes: bytes,
    mime: str,
    tenant_id: int | None,
    db,  # AsyncSession
) -> dict:
    """
    Returns:
      {
        "mapping": {"title": "Col A", "start_date": "Col B", ...},  # null si no detecta
        "headers": [...],
        "source": "cache" | "ai" | "alias",
        "usage": {"used": N, "limit": N | None, "plan": str}
      }
    Raises ValueError on rate limit exceeded.
    """
    from datetime import datetime, timezone
    from sqlalchemy import select, func as sqlfunc
    from app.models.ai_mapping_cache import AiMappingCache
    from app.models.tenant import Tenant
    from app.models.plan import Plan
    from app.core.config import settings

    # ── Extraer headers y sample del archivo ──────────────────────────────────
    if is_msproject_xml(file_bytes):
        raise ValueError("MS Project XML no necesita detección de columnas.")

    if mime == "text/csv" or mime == "application/csv":
        headers, data_rows = _rows_from_csv(file_bytes)
    else:
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        headers, data_rows = _rows_from_sheet(wb.active)

    headers = [str(h).strip() for h in headers if str(h).strip()]
    if not headers:
        raise ValueError("No se encontraron encabezados en el archivo.")

    sample_rows = data_rows[:3]
    h_hash = _headers_hash(headers)

    # ── Determinar plan y límite ──────────────────────────────────────────────
    plan_name = "sin_plan"
    monthly_limit: int | None = _DEFAULT_LIMIT

    if tenant_id is not None:
        tenant = await db.get(Tenant, tenant_id)
        if tenant and tenant.plan_id:
            plan = await db.get(Plan, tenant.plan_id)
            if plan:
                plan_name = plan.name
                monthly_limit = _AI_MONTHLY_LIMITS.get(plan.name, _DEFAULT_LIMIT)

    # ── Verificar caché ──────────────────────────────────────────────────────
    cached = (await db.execute(
        select(AiMappingCache).where(
            AiMappingCache.tenant_id == tenant_id,
            AiMappingCache.headers_hash == h_hash,
        )
    )).scalar_one_or_none()

    if cached:
        mapping = _json.loads(cached.mapping)
        # Contar uso del mes para informar al frontend
        month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        used = (await db.execute(
            select(sqlfunc.count()).where(
                AiMappingCache.tenant_id == tenant_id,
                AiMappingCache.created_at >= month_start,
            )
        )).scalar_one()
        return {
            "mapping": mapping,
            "headers": headers,
            "source": "cache",
            "usage": {"used": used, "limit": monthly_limit, "plan": plan_name},
        }

    # ── Verificar rate limit ──────────────────────────────────────────────────
    if monthly_limit is not None:
        month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        used_this_month = (await db.execute(
            select(sqlfunc.count()).where(
                AiMappingCache.tenant_id == tenant_id,
                AiMappingCache.created_at >= month_start,
            )
        )).scalar_one()
        if used_this_month >= monthly_limit:
            raise ValueError(
                f"Límite de {monthly_limit} detecciones IA por mes alcanzado "
                f"(plan {plan_name}). Remapeá las columnas manualmente."
            )
    else:
        used_this_month = 0

    # ── Intentar detección por alias primero (gratis) ─────────────────────────
    alias_map = _detect_column_map(headers)
    alias_result = {field: headers[idx] if idx is not None else None for field, idx in alias_map.items()}
    # Si detecta título y al menos una fecha, no necesitamos IA
    if alias_result.get("title") and (alias_result.get("start_date") or alias_result.get("due_date")):
        return {
            "mapping": alias_result,
            "headers": headers,
            "source": "alias",
            "usage": {"used": used_this_month, "limit": monthly_limit, "plan": plan_name},
        }

    # ── Llamar a Haiku ────────────────────────────────────────────────────────
    if not settings.ANTHROPIC_API_KEY:
        # Sin key → retornar detección parcial de aliases
        return {
            "mapping": alias_result,
            "headers": headers,
            "source": "alias",
            "usage": {"used": used_this_month, "limit": monthly_limit, "plan": plan_name},
        }

    sample_lines = []
    for i, row in enumerate(sample_rows):
        parts = [f"{h}: {str(row[j]).strip() if j < len(row) else ''}" for j, h in enumerate(headers)]
        sample_lines.append(f"Fila {i+1}: " + " | ".join(parts))

    prompt = (
        "Sos un experto en interpretar planillas de construcción/obra. Tu trabajo es identificar qué columna "
        "corresponde a cada campo de una tarea de obra, aunque los nombres sean abreviados, inusuales o estén "
        "en otro idioma.\n\n"
        f"ENCABEZADOS DISPONIBLES: {_json.dumps(headers, ensure_ascii=False)}\n\n"
        "DATOS DE EJEMPLO (primeras filas):\n" + "\n".join(sample_lines) + "\n\n"
        "CAMPOS A MAPEAR:\n"
        "- title: Nombre o título de la tarea. Buscá: 'Item', 'Nombre', 'Task', 'Actividad', 'Tarea', 'Descripción', "
        "o cualquier columna cuyo contenido sea texto descriptivo de la actividad.\n"
        "- start_date: Fecha de inicio. Buscá: 'Start', 'Inicio', 'Comienzo', 'Start?', 'Fecha inicio', o similares.\n"
        "- due_date: Fecha de fin/vencimiento. Buscá: 'Fin', 'Finish', 'End', 'Fecha cierre', 'Vencimiento', 'Fecha fin', o similares.\n"
        "- responsible: Responsable o asignado. Buscá: 'RESP.', 'Resp', 'Responsable', 'Recurso', 'Asignado', 'Resource', o similares.\n"
        "- predecessors: Predecesoras/dependencias (números). Buscá: 'dep.', 'Dep', 'Predecesoras', 'Pred.', 'Depends', o similares.\n\n"
        "REGLAS CRÍTICAS:\n"
        "1. Usá el nombre EXACTO del encabezado tal como aparece en la lista.\n"
        "2. SIEMPRE elegí la mejor opción disponible aunque el nombre sea abreviado o poco obvio. "
        "Es mejor asignar la columna más probable que dejarla en null.\n"
        "3. Para 'title': si hay una columna con texto libre de actividades (aunque se llame 'Item', 'Ítem', etc.), asignala.\n"
        "4. Solo usá null si estás SEGURO de que ninguna columna puede corresponder a ese campo.\n"
        "5. Mirá los datos de ejemplo para entender el contenido real de cada columna.\n\n"
        "Respondé ÚNICAMENTE con JSON válido, sin texto adicional:\n"
        '{"title": "...", "start_date": "...", "due_date": "...", "responsible": "...", "predecessors": "..."}'
    )

    import anthropic
    client = anthropic.AsyncAnthropic(api_key=settings.ANTHROPIC_API_KEY)
    response = await client.messages.create(
        model=settings.CLAUDE_MODEL,
        max_tokens=256,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = response.content[0].text.strip()

    # Parsear JSON de la respuesta
    try:
        # Extraer JSON si viene con texto extra
        start = raw.index("{")
        end = raw.rindex("}") + 1
        mapping = _json.loads(raw[start:end])
    except (ValueError, _json.JSONDecodeError):
        mapping = alias_result  # fallback a aliases

    # Asegurar que solo tenga los campos esperados
    mapping = {k: mapping.get(k) for k in _FIELD_DESCRIPTIONS}

    # ── Guardar en caché ──────────────────────────────────────────────────────
    cache_row = AiMappingCache(
        tenant_id=tenant_id,
        headers_hash=h_hash,
        mapping=_json.dumps(mapping, ensure_ascii=False),
    )
    db.add(cache_row)
    # El commit lo hace get_db()

    return {
        "mapping": mapping,
        "headers": headers,
        "source": "ai",
        "usage": {"used": used_this_month + 1, "limit": monthly_limit, "plan": plan_name},
    }
